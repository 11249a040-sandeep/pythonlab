import re, os, zipfile, PyPDF2, sys, subprocess

def install_missing_packages():
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "PyPDF2"])
install_missing_packages()
import openpyxl

def exp1():
    print("\n=== 1. Calculation of Test Average ===")
    a, b, c = map(float, input("Enter three test marks: ").split())
    avg = (a + b + c - min(a, b, c)) / 2
    print("Best of two test average:", avg)

def exp2():
    print("\n=== 2. Palindrome Check & Digit Occurrence Count ===")
    n = input("Enter a number: ")
    print("Palindrome" if n == n[::-1] else "Not Palindrome")
    for d in sorted(set(n)):
        print(d, ":", n.count(d))

def exp3():
    print("\n=== 3. Fibonacci Sequence ===")
    n = int(input("Enter N (>0): "))
    if n <= 0:
        print("Invalid input")
        return
    f, s = 0, 1
    print(f, s, end=" ")
    for _ in range(2, n):
        f, s = s, f + s
        print(s, end=" ")
    print()

def exp4a():
    print("\n=== 4A. Binary to Decimal Conversion ===")
    b = input("Enter binary number: ")
    print("Decimal:", int(b, 2))

def exp4b():
    print("\n=== 4B. Octal to Hexadecimal Conversion ===")
    o = input("Enter octal number: ")
    print("Hexadecimal:", hex(int(o, 8))[2:].upper())

def exp5():
    print("\n=== 5. Sentence Statistics ===")
    s = input("Enter a sentence: ")
    words = len(s.split())
    digits = sum(c.isdigit() for c in s)
    upper = sum(c.isupper() for c in s)
    lower = sum(c.islower() for c in s)
    print("Words:", words, "Digits:", digits, "Uppercase:", upper, "Lowercase:", lower)

def exp6():
    print("\n=== 6. String Similarity ===")
    a = input("Enter first string: ")
    b = input("Enter second string: ")
    from difflib import SequenceMatcher
    print("Similarity Ratio:", round(SequenceMatcher(None, a, b).ratio(), 3))

def exp7():
    print("\n=== 7. Insertion Sort & Merge Sort on Lists ===")
    lst = list(map(int, input("Enter numbers: ").split()))
    for i in range(1, len(lst)):
        key = lst[i]
        j = i - 1
        while j >= 0 and lst[j] > key:
            lst[j + 1] = lst[j]
            j -= 1
        lst[j + 1] = key
    print("Insertion Sort Result:", lst)
    def merge_sort(arr):
        if len(arr) <= 1:
            return arr
        mid = len(arr)//2
        left = merge_sort(arr[:mid])
        right = merge_sort(arr[mid:])
        res = []
        i = j = 0
        while i < len(left) and j < len(right):
            if left[i] < right[j]:
                res.append(left[i])
                i += 1
            else:
                res.append(right[j])
                j += 1
        res.extend(left[i:])
        res.extend(right[j:])
        return res
    print("Merge Sort Result:", merge_sort(lst))

def exp8():
    print("\n=== 8. Check Phone Number ===")
    def isphonenumber(num):
        return len(num) == 12 and num[:3].isdigit() and num[3] == '-' and num[4:7].isdigit() and num[7] == '-' and num[8:].isdigit()
    s = input("Enter number (415-555-4242): ")
    print("Manual check:", "Valid" if isphonenumber(s) else "Invalid")
    p = re.compile(r'^\d{3}-\d{3}-\d{4}$')
    print("Regex check:", "Valid" if p.match(s) else "Invalid")

def exp9a():
    print("\n=== 9A. Search Phone Numbers in File ===")
    f = input("Enter file name: ")
    with open(f, encoding='utf-8') as file:
        text = file.read()
    pattern = r'(\+91[6-9]\d{9}|\b[6-9]\d{9}\b|0[6-9]\d{9})'
    found = re.findall(pattern, text)
    if found:
        print("Phone Numbers Found:")
        for n in sorted(set(found)):
            print(n)
    else:
        print("No phone numbers found")

def exp9b():
    print("\n=== 9B. Search Emails in File ===")
    f = input("Enter file name: ")
    with open(f, encoding='utf-8') as file:
        text = file.read()
    pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    found = re.findall(pattern, text)
    if found:
        print("Emails Found:")
        for e in sorted(set(found)):
            print(e)
    else:
        print("No emails found")

def exp10():
    print("\n=== 10. File Operations ===")
    f = input("Enter file name: ")
    n = int(input("Enter N: "))
    w = input("Enter word: ")
    with open(f, encoding='utf-8') as file:
        lines = file.readlines()
    print("First", n, "lines:")
    print("".join(lines[:n]))
    print("Word frequency:", sum(line.count(w) for line in lines))

def exp11():
    print("\n=== 11. Zip Operation on a Folder ===")
    folder = input("Enter folder name: ")
    with zipfile.ZipFile(folder + '.zip', 'w') as z:
        for root, dirs, files in os.walk(folder):
            for file in files:
                z.write(os.path.join(root, file))
    print("Folder backed up as", folder + '.zip')

def exp12():
    print("\n=== 12. Inheritance - Area Calculation ===")
    import math
    class Shape: pass
    class Triangle(Shape):
        def area(self, b, h): return 0.5 * b * h
    class Circle(Shape):
        def area(self, r): return math.pi * r * r
    class Rectangle(Shape):
        def area(self, l, b): return l * b
    t, c, r = Triangle(), Circle(), Rectangle()
    print("Triangle area:", t.area(10, 5))
    print("Circle area:", c.area(5))
    print("Rectangle area:", r.area(4, 6))

def exp13():
    print("\n=== 13. Employee Details ===")
    class Employee:
        def __init__(self, n, i, d, s):
            self.name, self.id, self.dept, self.salary = n, i, d, s
        def update(self, dep, inc):
            if self.dept == dep:
                self.salary += inc
    n = int(input("Enter number of employees: "))
    emp = [Employee(input("Name: "), input("ID: "), input("Dept: "), float(input("Salary: "))) for _ in range(n)]
    dep = input("Enter department to update: ")
    inc = float(input("Enter increment: "))
    for e in emp: e.update(dep, inc)
    for e in emp: print(e.name, e.id, e.dept, e.salary)

def exp14():
    print("\n=== 14. Polymorphism and Inheritance - Palindrome Check ===")
    class Palindrome:
        def check(self, x):
            s = str(x)
            return s == s[::-1]
    p = Palindrome()
    s = input("Enter string or number: ")
    print("Palindrome" if p.check(s) else "Not Palindrome")

def exp15():
    print("\n=== 15. Spreadsheet Operations ===")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Amit", 21])
    ws.append(["Sita", 22])
    wb.save("data.xlsx")
    wb2 = openpyxl.load_workbook("data.xlsx")
    ws2 = wb2.active
    for row in ws2.iter_rows(values_only=True):
        print(row)

def exp16():
    print("\n=== 16. Merge Selected Pages from Multiple PDFs ===")
    merger = PyPDF2.PdfMerger()
    n = int(input("Enter number of PDFs: "))
    for _ in range(n):
        f = input("Enter PDF filename: ")
        s, e = map(int, input("Enter start and end page: ").split())
        merger.append(f, pages=(s-1, e))
    merger.write("merged.pdf")
    merger.close()
    print("Merged as merged.pdf")

experiments = {
    1: exp1, 2: exp2, 3: exp3, 4.1: exp4a, 4.2: exp4b, 5: exp5, 6: exp6, 7: exp7,
    8: exp8, 9.1: exp9a, 9.2: exp9b, 10: exp10, 11: exp11, 12: exp12,
    13: exp13, 14: exp14, 15: exp15, 16: exp16
}

def show_menu():
    print("\n====== PYTHON LAB EXPERIMENTS ======")
    print("1.  Calculation of Test Average")
    print("2.  Palindrome Check & Digit Occurrence Count")
    print("3.  Fibonacci Sequence")
    print("4.1 Binary to Decimal Conversion")
    print("4.2 Octal to Hexadecimal Conversion")
    print("5.  Sentence Statistics")
    print("6.  String Similarity")
    print("7.  Insertion Sort & Merge Sort")
    print("8.  Check Phone Number")
    print("9.1 Search Phone Numbers in File")
    print("9.2 Search Emails in File")
    print("10. File Operations")
    print("11. Zip Operation on a Folder")
    print("12. Inheritance - Area Calculation")
    print("13. Employee Details")
    print("14. Polymorphism & Inheritance - Palindrome")
    print("15. Spreadsheet Operations")
    print("16. Merge Selected Pages from Multiple PDFs")

while True:
    show_menu()
    ch = input("\nEnter experiment number (or 'exit' to quit): ")
    if ch.lower() == 'exit':
        break
    try:
        ch = float(ch)
        if ch in experiments:
            experiments[ch]()
        else:
            print("Invalid choice")
    except:
        print("Invalid input")
